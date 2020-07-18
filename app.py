from flask import Flask
from flask import request
from flask import render_template
import folium
from folium import plugins
from folium.plugins import HeatMap
from openpyxl import *
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
from geopy.geocoders import Nominatim
import numpy as np



wb = load_workbook('data.xlsx')
ws=wb["Sheet1"]

app = Flask(__name__)
geolocator = Nominatim()


@app.route("/")
def index():
   return render_template("index.html")


@app.route("/index.html")
def index1():
   return render_template("index.html")


@app.route("/form.html")
def form():
    return render_template("form.html")

@app.route("/map.html")
def map():
    folium_map = folium.Map(location=[40.745857,-74.025598],
                        zoom_start=13)
    mapdata=pd.read_excel('data.xlsx')
    heat_df = mapdata[['Latitude', 'Longitude']]
    heat_df = heat_df.dropna(axis=0, subset=['Latitude','Longitude'])
    heatdata = mapdata[['Latitude', 'Longitude']].to_numpy()
    folium_map.add_children(plugins.HeatMap(heatdata, radius=20))
    return folium_map._repr_html_()

@app.route("/", methods=["POST"])
def getvalue():
    street= request.form["Address"]
    town= request.form["Town"]
    state= request.form["State"]
    address = street + ", " + town + ", " + state
    
    '''
    address = request.form["address"]
    '''
    location = geolocator.geocode(address)
    df=pd.read_excel('data.xlsx')
    wcell1=ws.cell(len(df["Longitude"])+2,2)
    wcell1.value=location.longitude
    wcell2=ws.cell(len(df["Latitude"])+2,3)
    wcell2.value=location.latitude
    wcell3=ws.cell(len(df["Address"])+2,4)
    wcell3.value=location.address
    wb.save('data.xlsx')
    folium_map = folium.Map(location=[location.latitude, location.longitude],
                        zoom_start=17)
    
    
    mapdata = pd.read_excel('data.xlsx')
    heat_df = mapdata[['Latitude', 'Longitude']]
    heat_df = heat_df.dropna(axis=0, subset=['Latitude','Longitude'])

    # List comprehension to make out list of lists
    '''
    heat_data = [[row['Latitude'],row['Longitude']] for index, row in heat_df.iterrows()]
    '''
    
    heatdata = mapdata[['Latitude', 'Longitude']].to_numpy()
    # Plot it on the map
    '''
    HeatMap(heat_data).add_to(folium_map)
    '''
    folium_map.add_children(plugins.HeatMap(heatdata, radius=20))

    return folium_map._repr_html_()

if __name__ == "__main__":
    app.run()
  